import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook


def biotnum(h_convec, ro, k_medium):
    """
    Biot's Number function
    :param h_convec: This is the convection term that will come from your choice, i.e. air or water
    :param ro: the radius of the object, or width depending on the application
    :param k_medium: the thermal conductivity of the object itself.
    :return: The Biot's Number
    """
    biots_list = []
    for i in range(len(h_convec)):
        biots_num = h_convec[i] * ro / k_medium
        biots_list.append(biots_num)

    return biots_list


def excel2py(biots_num):
    """
    This function will mass extract the constants needed for a given scenario of the common geometry used
    ####### CYLINDER IS CURRENTLY IN USE. PLEASE UPDATE YOUR OWN EXCEL SPREADSHEET AS NEEDED FOR MORE GEOMETRY
    :param biots_num: The massive list of data points just made
    :return: A1 and Lambda-1 values needed
    """

    biots_conv = load_workbook("E:\\Excel and Py\\TensorVals.xlsx")

    sheet_needed = input("What is the sheet you need for this data? Enter the EXACT name found for the sheet on your"
                         "Excel document...: ")
    last_question = int(input("What cell should we begin from? "))

    if sheet_needed in biots_conv.sheetnames:
        desired_sheet = biots_conv[sheet_needed]

        cells = []
        cellm = []
        celll = []
        cellinf = []
        for cell_extractor in desired_sheet[last_question]:
            cell_r = cell_extractor.value
            cells.append(cell_r)
        for cell_extractor in desired_sheet[last_question + 1]:
            cell_r = cell_extractor.value
            cellm.append(cell_r)
        for cell_extractor in desired_sheet[last_question + 2]:
            cell_r = cell_extractor.value
            celll.append(cell_r)
        for cell_extractor in desired_sheet[3]:
            cell_r = cell_extractor.value
            cellinf.append(cell_r)

        print(cells, cellm, celll)

    interp_value = []
    for i in range(len(biots_num)):
        if biots_num[i] >= cells[0] and biots_num[i] <= cellm[0]:

            lambda1_inter = cells[1] + (biots_num[i] - cells[0]) * ((cellm[1] - cells[1]) / (cellm[0] - cells[0]))

            A1_interp = cells[2] + (biots_num[i] - cells[0]) * ((cellm[2] - cells[2]) / (cellm[0] - cells[0]))

            interp_value.append((lambda1_inter, A1_interp))
        elif biots_num[i] >= cellm[0] and biots_num[i] <= celll[0]:
            lambda1_inter = cellm[1] + (biots_num[i] - cellm[0]) * ((celll[1] - cellm[1]) / (celll[0] - cellm[0]))

            A1_interp = cellm[2] + (biots_num[i] - cellm[0]) * ((celll[2] - cellm[2]) / (celll[0] - cellm[0]))

            interp_value.append((lambda1_inter, A1_interp))
        else:
            lambda1_inter = celll[1] + (biots_num[i] - celll[0]) * ((cellinf[1] - celll[1]) / (cellinf[0] - celll[0]))

            A1_interp = celll[2] + (biots_num[i] - celll[0]) * ((cellinf[2] - celll[2]) / (cellinf[0] - celll[0]))

            interp_value.append((lambda1_inter, A1_interp))

    print(interp_value)
    return interp_value


def time_finder(a1_lambda1):
    """
    This will find the time finally. Just put in your own T values
    :param a1_lambda1: Constants for the equation
    :return: The time finally needed
    """
    time_storage = []
    for i in range(len(a1_lambda1)):
        tau = (np.log((((102 - 500) / (20 - 500))/a1_lambda1[i][0]))) / (-(a1_lambda1[i][1]**2))


        t = tau /(4.27 * 10 **-4)

        time_storage.append(t/60)

    print(time_storage)

    return time_storage


def money_calc(times):
    """
    This will just calculate the cost over time.
    :param times:
    :return:
    """
    total_cost = []
    total_cost_2 = []
    for i in range(len(times_container[0])):
        total_cost.append(((0.95*1224000)/(100000*24*60)) * times_container[0][i])
        total_cost_2.append(((0.95*1224000)/(100000*24*60)) * times_container[1][i])

    print(total_cost)
    return total_cost, total_cost_2


def grapher(x, y):
    """
    Graphing our solution at the end
    :param x:
    :param y:
    :return:
    """
    plt.plot(h_linspace, times_container[0], "-b", label="Stainless Steel-304")
    plt.plot(h_linspace, times_container[1], "-r", label="Cast Iron")
    plt.xlabel("Coefficient of Convection for Water (W/m^2 * K)")
    plt.ylabel("Time to reach boil (Minutes)")
    plt.title("Convection vs. Time to Reach Boil for Saltwater")
    plt.legend()
    plt.xlim(0, 100)
    plt.ylim(0, 50)
    plt.show()

    plt.plot(h_linspace, times_container[2], "-b", label="Stainless Steel-304")
    plt.plot(h_linspace, times_container[3], "-r", label="Cast Iron")
    plt.xlabel("Coefficient of Convection for Water (W/m^2 * K)")
    plt.ylabel("Time to reach boil (Minutes)")
    plt.title("Convection vs. Time to Reach Boil for Saltwater with Half of the Radius")
    plt.legend()
    plt.xlim(0, 100)
    plt.ylim(0, 50)
    plt.show()

    cost = money_calc(times_container[0])

    plt.plot(times_container[0], cost[0], color="blue", label="Stainless Steel-304")
    plt.xlabel("Time to Reach Boil per Batch (Minutes)")
    plt.ylabel("Cost Associated with Time to Reach Boil for Saltwater (Dollars)")
    plt.title("Cost vs. Time to Reach Boil")
    plt.legend()
    plt.xlim(0, 28)
    plt.ylim(0, 1)
    plt.show()


if __name__ == "__main__":
    times_container = []
    welcome_opt = input("Welcome! Is this a simple calculation, or advanced analysis of some sort? (s/a to continue): ")

    if welcome_opt == "a":
        concluding_msg = 0
        while concluding_msg != "stop":
            opt_1 = input("What is changing to conduct this analysis: h, ro/w, k, or something else?: ")
            if opt_1 == "h":
                ro = float(input("Please input the value of your radius or width: "))
                k = float(input("Please input the value of your thermal conductivity: "))
                h_low = float(input("What is the lowest possible convection value?: "))
                h_high = float(input("What is the highest possible convection value?: "))
                h_linspace = np.linspace(h_low, h_high, num=1250)

                biot_needed = biotnum(h_linspace, ro, k)

                print(biot_needed)

                a1_lambda1 = excel2py(biot_needed)

                times = time_finder(a1_lambda1)


            times_container.append(times)

            concluding_msg = input("Enter 'Stop' when you are ready to graph... ")




    graphs = grapher(h_linspace, times_container)
    print("Exiting program... please wait.")




