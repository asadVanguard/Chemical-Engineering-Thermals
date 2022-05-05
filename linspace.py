import numpy as np
import openpyxl
from openpyxl import load_workbook

comparator_val = 75

tense = load_workbook("E:\\Excel and Py\\TensorVals.xlsx")

needed_sheet = "Sheet2"

if needed_sheet in tense.sheetnames:
    tense_opened = tense.active
    desired_sheet = tense[needed_sheet]





    cells = []
    for c1 in desired_sheet[2]:
        cell_r = c1.value

        cells.append(cell_r)



    print(cells)